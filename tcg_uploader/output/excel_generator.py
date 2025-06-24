"""Excel file generation for eBay upload"""

import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import asyncio

from ..models import CardData
from ..config import Config
from ..api.openai_titles import OpenAITitleOptimizer
from .ebay_formatter import EbayFormatter
from ..utils.logger import logger

class ExcelGenerator:
    def __init__(self, config: Config):
        self.config = config
        self.output_folder = config.output_folder
        self.ebay_formatter = EbayFormatter(config)
        
        # Initialize OpenAI if available
        if config.openai_api_key:
            self.title_optimizer = OpenAITitleOptimizer(config.openai_api_key)
        else:
            self.title_optimizer = None
    
    async def create_excel(self, cards: List[CardData]) -> str:
        """Create Excel file with eBay listings"""
        # Optimize titles
        if self.title_optimizer:
            card_dicts = [card.to_dict() for card in cards]
            optimized_titles = await self.title_optimizer.optimize_titles(card_dicts)
        else:
            optimized_titles = [self._generate_fallback_title(card) for card in cards]
        
        # Create eBay rows
        ebay_rows = []
        for i, card in enumerate(cards):
            title = optimized_titles[i] if i < len(optimized_titles) else self._generate_fallback_title(card)
            ebay_row = self.ebay_formatter.format_card(card, title, i)
            ebay_rows.append(ebay_row)
        
        # Create DataFrame
        df = pd.DataFrame(ebay_rows)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'tcg_listings_{timestamp}.xlsx'
        output_path = self.output_folder / filename
        
        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='eBay_Upload', index=False)
            
            # Apply conditional formatting
            self._apply_formatting(writer.book['eBay_Upload'], df)
        
        return str(output_path)
    
    def _generate_fallback_title(self, card: CardData) -> str:
        """Generate fallback title"""
        parts = []
        
        # Game prefix
        if card.game.lower() in ['pokémon', 'pokemon']:
            parts.append('Pokémon')
        elif card.game.lower() in ['magic: the gathering', 'mtg']:
            parts.append('MTG')
        else:
            parts.append(card.game)
        
        # Card details
        parts.append(card.name)
        
        if card.number:
            parts.append(card.number)
        
        if card.set_name:
            parts.append(card.set_name)
        
        # Rarity/finish if meaningful
        if card.rarity and card.rarity.lower() not in ['common', 'normal']:
            parts.append(card.rarity)
        
        if card.finish and card.finish.lower() not in ['normal', 'regular']:
            parts.append(card.finish)
        
        # Unique characteristics
        for char in card.unique_characteristics[:1]:
            parts.append(char)
            break
        
        # Condition
        parts.append('NM/LP')
        
        # Language
        if card.language.lower() in ['japanese', 'jp']:
            parts.append('JP')
        
        # Join and truncate
        title = ' '.join(parts)
        if len(title) > 80:
            title = title[:77] + "..."
        
        return title
    
    def _apply_formatting(self, worksheet, df: pd.DataFrame):
        """Apply conditional formatting to worksheet"""
        from openpyxl.styles import PatternFill
        
        # Find ReviewFlag column
        review_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'ReviewFlag':
                review_col = idx
                break
        
        if not review_col:
            return
        
        # Define fills
        green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        
        # Apply formatting
        for row_num in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_num, column=review_col)
            
            if cell.value == 'OK':
                fill = green_fill
            else:
                fill = red_fill
            
            # Apply to entire row
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_num, column=col).fill = fill