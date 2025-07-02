"""Excel file generation for eBay upload with variation support"""

import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import asyncio

from ..models import CardData
from ..config import Config
from ..api.openai_titles import OpenAITitleOptimizer
from .ebay_formatter import EbayFormatter
from ..utils.logger import logger

class ExcelGenerator:
    def __init__(self, config: Config):
        self.config = config
        self.output_folder = config.output_folder
        self.ebay_formatter = EbayFormatter(config)
        
        # Initialize OpenAI if available
        if config.openai_api_key:
            self.title_optimizer = OpenAITitleOptimizer(config.openai_api_key)
        else:
            self.title_optimizer = None
    
    async def create_excel(self, cards: List[CardData]) -> str:
        """Create Excel file with eBay listings including variation support"""
        # Optimize titles
        if self.title_optimizer:
            card_dicts = [card.to_dict() for card in cards]
            optimized_titles = await self.title_optimizer.optimize_titles(card_dicts)
        else:
            optimized_titles = [self._generate_fallback_title(card) for card in cards]
        
        # Create eBay rows with variation support
        ebay_rows = []
        row_index = 0
        
        for i, card in enumerate(cards):
            title = optimized_titles[i] if i < len(optimized_titles) else self._generate_fallback_title(card)
            
            if card.is_variation_parent and len(card.variations) > 1:
                # Create parent row and variation rows
                parent_rows = self._create_variation_listing(card, title, row_index)
                ebay_rows.extend(parent_rows)
                row_index += len(parent_rows)
            else:
                # Single listing (no variations)
                ebay_row = self.ebay_formatter.format_card(card, title, row_index)
                ebay_rows.append(ebay_row)
                row_index += 1
        
        # Create DataFrame
        df = pd.DataFrame(ebay_rows)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'tcg_listings_{timestamp}.xlsx'
        output_path = self.output_folder / filename
        
        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='eBay_Upload', index=False)
            
            # Apply conditional formatting
            self._apply_formatting(writer.book['eBay_Upload'], df)
        
        logger.info(f"📊 Created Excel with {len(cards)} listings ({len(ebay_rows)} total rows)")
        
        return str(output_path)
    
    def _create_variation_listing(self, card: CardData, title: str, start_index: int) -> List[Dict[str, Any]]:
        """Create parent and child rows for variation listing"""
        rows = []
        
        # Create parent row
        parent_row = self.ebay_formatter.format_card(card, title, start_index)
        
        # Modify parent row for variations
        parent_row['*Relationship'] = 'Parent'
        parent_row['*Quantity'] = ''  # Parent has no quantity
        parent_row['*StartPrice'] = ''  # Parent has no price
        parent_row['Variation Theme'] = 'Copy Number'  # Our variation theme
        
        # Remove variation-specific fields from parent
        parent_row['CustomLabel'] = card.variations[0]['custom_label'] if card.variations else parent_row['CustomLabel']
        
        rows.append(parent_row)
        
        # Create child rows for each variation
        for idx, variation in enumerate(card.variations):
            child_row = self._create_variation_row(card, title, variation, variation.get('copy_number', idx + 1))
            rows.append(child_row)
        
        return rows
    
    def _create_variation_row(self, parent_card: CardData, title: str, 
                             variation: Dict[str, Any], copy_number: int) -> Dict[str, Any]:
        """Create a child row for a variation"""
        # Start with a copy of parent data
        child_row = {}
        
        # Required variation fields
        child_row['*Action(SiteID=US|Country=US|Currency=USD|Version=1193)'] = ''  # Blank for child
        child_row['*Relationship'] = 'Variation'
        child_row['*VariationSpecifics'] = f'Copy Number=Copy #{copy_number}'
        child_row['CustomLabel'] = variation['custom_label']
        
        # Variation-specific data
        child_row['*Title'] = ''  # Child rows don't need title
        child_row['*Quantity'] = 1
        child_row['*StartPrice'] = parent_card.final_price
        
        # Images - each variation gets its own
        if variation['image_urls']:
            urls_with_cache_bust = [f"{url}?cache-bust={i}" for i, url in enumerate(variation['image_urls'])]
            child_row['PicURL'] = "|".join(urls_with_cache_bust)
        else:
            child_row['PicURL'] = ''
        
        # All other fields inherit from parent (empty in child rows)
        child_row['*Category'] = ''
        child_row['*ConditionID'] = ''
        child_row['*Description'] = ''
        child_row['*Format'] = ''
        child_row['*Duration'] = ''
        child_row['*Location'] = ''
        child_row['PostalCode'] = ''
        child_row['*DispatchTimeMax'] = ''
        
        # Business policies (inherit from parent)
        child_row['PaymentProfileName'] = ''
        child_row['ShippingProfileName'] = ''
        child_row['ReturnProfileName'] = ''
        
        # Item specifics (all blank for child rows - inherited from parent)
        for field in ['*C:Game', 'C:Card Name', 'C:Character', 'C:Set', 'C:Rarity', 
                      'C:Card Number', 'C:Graded', 'C:Manufacturer', 'C:Autographed', 
                      'C:Language', 'C:Finish', 'C:Features', 'C:Card Size', 
                      'C:Year Manufactured', 'C:Vintage', 'C:Country/Region of Manufacture',
                      'C:Defense/Toughness', 'C:HP', 'C:Card Type', 'C:Attribute/MTG:Color',
                      'CD:Card Condition - (ID: 40001)', 'StoreCategory', 'Subtitle']:
            child_row[field] = ''
        
        # Tracking columns
        child_row['ConfidenceScore'] = ''
        child_row['ReviewFlag'] = ''
        child_row['PriceSource'] = ''
        child_row['TCGPlayerLink'] = ''
        child_row['ProcessingNotes'] = f'Variation copy #{copy_number}'
        child_row['ImageCount'] = len(variation['image_urls'])
        
        # Add variation-specific column
        child_row['C:Copy Number'] = f'Copy #{copy_number}'
        
        return child_row
    
    def _generate_fallback_title(self, card: CardData) -> str:
        """Generate fallback title with smart filtering of boring terms"""
        parts = []
        
        # Game prefix
        if card.game.lower() in ['pokémon', 'pokemon']:
            parts.append('Pokémon')
        elif card.game.lower() in ['magic: the gathering', 'mtg']:
            parts.append('MTG')
        else:
            parts.append(card.game)
        
        # Card details
        parts.append(card.name)
        
        if card.number:
            parts.append(card.number)
        
        if card.set_name:
            parts.append(card.set_name)
        
        # Rarity - ONLY if attractive (skip common/uncommon)
        if card.rarity and card.rarity.lower() not in ['common', 'uncommon', 'normal', 'regular']:
            parts.append(card.rarity)
        
        # Finish - ONLY if meaningful (skip boring finishes)
        if card.finish and card.finish.lower() not in ['normal', 'regular', 'standard', 'non-holo', 'non-foil']:
            parts.append(card.finish)
        
        # Unique characteristics - skip Unlimited (it's the default)
        for char in card.unique_characteristics[:1]:
            if char.lower() != 'unlimited':
                parts.append(char)
                break
        
        # Condition
        parts.append('NM/LP')
        
        # Language
        if card.language.lower() in ['japanese', 'jp']:
            parts.append('JP')
        
        # Add variation indicator if applicable
        if card.is_variation_parent and card.variation_count > 1:
            parts.append(f'({card.variation_count} Available)')
        
        # Join and truncate
        title = ' '.join(parts)
        if len(title) > 80:
            title = title[:77] + "..."
        
        return title
    
    def _apply_formatting(self, worksheet, df: pd.DataFrame):
        """Apply conditional formatting to worksheet"""
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Find columns
        review_col = None
        relationship_col = None
        
        for idx, col in enumerate(df.columns, 1):
            if col == 'ReviewFlag':
                review_col = idx
            elif col == '*Relationship':
                relationship_col = idx
        
        # Define fills
        green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        parent_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        variation_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
        
        # Apply formatting
        for row_num in range(2, len(df) + 2):
            # Relationship-based formatting
            if relationship_col:
                rel_cell = worksheet.cell(row=row_num, column=relationship_col)
                if rel_cell.value == 'Parent':
                    # Bold parent rows
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_num, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = parent_fill
                elif rel_cell.value == 'Variation':
                    # Light purple for variations
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_num, column=col).fill = variation_fill
            
            # Review flag formatting (only for non-variation rows)
            if review_col:
                review_cell = worksheet.cell(row=row_num, column=review_col)
                if review_cell.value == 'OK':
                    fill = green_fill
                elif review_cell.value and review_cell.value != '':
                    fill = red_fill
                else:
                    continue
                
                # Apply review formatting only to parent/single rows
                if not relationship_col or rel_cell.value != 'Variation':
                    for col in range(1, worksheet.max_column + 1):
                        if worksheet.cell(row=row_num, column=col).fill == parent_fill:
                            continue  # Don't override parent formatting
                        worksheet.cell(row=row_num, column=col).fill = fill